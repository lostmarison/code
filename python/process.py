from openpyxl import load_workbook, Workbook

# 读取表格
workbook = load_workbook('计科学生信息.xlsx')
sheet1 = workbook.active  # 或者使用 workbook['SheetName'] 来指定工作表
rows = sheet1.max_row  # 读取行数
cols = sheet1.max_column  # 读取列数

# 创建所需列表
name = []  # 创建存储名字的空列表
gender = []  # 存储性别的空列表
student_id = []  # 存储学号的空列表
dorm = []  # 存储寝室的空列表
dorm_bed = []  # 存储床位的空列表

for row in sheet1.iter_rows(values_only=True):
    if row[5] == 2322107012:
        name.append(row[3])
        gender.append('男' if row[0] == 'C14' else '女')
        student_id.append(row[4])
        dorm.append(row[1])
        dorm_bed.append(row[2])

# 写入表格
write_workbook = Workbook()  # 创建一个新的工作簿
write_sheet = write_workbook.active  # 获取当前活动的工作表
headers = ["姓名", "性别", "学号", "寝室", "床位"]  # 定义表头
write_sheet.append(headers)  # 将表头写入第一行

# 将数据写入工作表
for i in range(len(name)):
    write_sheet.append([name[i], gender[i], student_id[i], dorm[i], dorm_bed[i]])

'''
# 指定保存位置和文件名
save_path = 'D:\\code\\python\\Excel_process\\student_information\\计科2班学生信息.xlsx'
write_workbook.save(save_path)  # 保存工作簿
'''
